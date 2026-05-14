"""
RTL Arabic Excel builders for single-day, date-range, and monthly attendance matrices.
"""

from __future__ import annotations

import io
from calendar import monthrange
from datetime import date, timedelta
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _styles():
    hf = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill("solid", start_color="1a3a5c")
    pfill = PatternFill("solid", start_color="d4edda")
    afill = PatternFill("solid", start_color="fee2e2")
    lfill = PatternFill("solid", start_color="fff3cd")
    blank_fill = PatternFill("solid", start_color="f5f5f5")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rgt = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="cccccc")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    return hf, hfill, pfill, afill, lfill, blank_fill, ctr, rgt, brd


def _cell_fill_for_day_token(tok: str, pfill, afill, lfill, blank_fill):
    if tok == "ح":
        return pfill
    if tok == "غ":
        return afill
    if tok == "م":
        return lfill
    return blank_fill


def build_single_day_workbook(
    group_id: str,
    label: str,
    time_slot: str,
    attendance_date: date,
    student_names: List[str],
    records: Dict[str, str],
) -> bytes:
    """Legacy single-day sheet (م / اسم / حضور / ملاحظات)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "كشف الحضور"
    ws.sheet_view.rightToLeft = True
    hf, hfill, pfill, afill, _lf, _bf, ctr, rgt, brd = _styles()

    def is_present(raw):
        return raw in ("present", "حاضر", "متاخر")

    ws.merge_cells("A1:D1")
    ws["A1"] = f"كشف حضور - {label} | {time_slot} | تاريخ: {attendance_date.isoformat()}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1a3c5e")
    ws["A1"].alignment = ctr
    ws["A1"].fill = PatternFill("solid", start_color="e8f0fe")
    ws.row_dimensions[1].height = 30

    for col, h in enumerate(["م", "اسم الطالب", "الحضور", "ملاحظات"], 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hf
        c.fill = hfill
        c.alignment = ctr
        c.border = brd
    ws.row_dimensions[2].height = 25

    present_count = 0
    for i, name in enumerate(student_names, 1):
        row = i + 2
        raw_s = records.get(name, "")
        if not raw_s:
            raw_s = "غائب"
        status = "present" if is_present(raw_s) else "absent"
        if status == "present":
            present_count += 1
        fill = pfill if status == "present" else afill
        label_ar = "متأخر ⏰" if raw_s == "متاخر" else ("حاضر ✔" if status == "present" else "غائب ✘")
        for col, (val, aln) in enumerate(zip([i, name, label_ar, ""], [ctr, rgt, ctr, ctr]), 1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = aln
            c.border = brd
            c.font = Font(name="Arial", size=11)
            if col in (1, 3):
                c.fill = fill
        ws.row_dimensions[row].height = 22

    total = len(student_names)
    sr = total + 3
    ws.merge_cells(f"A{sr}:B{sr}")
    ws[f"A{sr}"] = f"إجمالي الحاضرين: {present_count} / {total}"
    ws[f"A{sr}"].font = Font(name="Arial", bold=True, size=11, color="155724")
    ws[f"A{sr}"].fill = pfill
    ws[f"A{sr}"].alignment = ctr
    ws[f"A{sr}"].border = brd
    ws.merge_cells(f"C{sr}:D{sr}")
    ws[f"C{sr}"] = f"إجمالي الغائبين: {total - present_count}"
    ws[f"C{sr}"].font = Font(name="Arial", bold=True, size=11, color="721c24")
    ws[f"C{sr}"].fill = afill
    ws[f"C{sr}"].alignment = ctr
    ws[f"C{sr}"].border = brd

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_matrix_workbook(
    title: str,
    subtitle: str,
    dates: List[date],
    student_names: List[str],
    matrix: List[List[str]],
    summaries: Dict[str, Tuple[int, int, int, float]],
    legend_extra: str = "",
) -> bytes:
    """
    Rows = students, columns = dates + summary (حاضر/غائب/متأخر/%).
    ``matrix`` aligns with ``student_names`` row order.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "الحضور"
    ws.sheet_view.rightToLeft = True
    hf, hfill, pfill, afill, lfill, blank_fill, ctr, rgt, brd = _styles()

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(6, 2 + len(dates) + 4))
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="1a3c5e")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].fill = PatternFill("solid", start_color="e8f0fe")
    ws["A1"].border = brd
    ws.row_dimensions[1].height = 36

    r2 = 2
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=max(6, 2 + len(dates) + 4))
    ws.cell(row=r2, column=1, value=subtitle + (" | " + legend_extra if legend_extra else ""))
    ws.cell(row=r2, column=1).font = Font(name="Arial", size=10, color="444444")
    ws.cell(row=r2, column=1).alignment = ctr
    ws.row_dimensions[r2].height = 28

    hdr_row = 3
    ws.cell(row=hdr_row, column=1, value="م").font = hf
    ws.cell(row=hdr_row, column=1).fill = hfill
    ws.cell(row=hdr_row, column=1).alignment = ctr
    ws.cell(row=hdr_row, column=1).border = brd
    ws.cell(row=hdr_row, column=2, value="اسم الطالب").font = hf
    ws.cell(row=hdr_row, column=2).fill = hfill
    ws.cell(row=hdr_row, column=2).alignment = ctr
    ws.cell(row=hdr_row, column=2).border = brd

    for j, d in enumerate(dates, start=1):
        col = 2 + j
        cell = ws.cell(row=hdr_row, column=col, value=d.strftime("%m-%d"))
        cell.font = hf
        cell.fill = hfill
        cell.alignment = ctr
        cell.border = brd

    sum_cols = ["حاضر", "غائب", "متأخر", "نسبة %"]
    base_col = 3 + len(dates)
    for k, lab in enumerate(sum_cols):
        c = ws.cell(row=hdr_row, column=base_col + k, value=lab)
        c.font = hf
        c.fill = hfill
        c.alignment = ctr
        c.border = brd

    ws.row_dimensions[hdr_row].height = 24

    for i, name in enumerate(student_names):
        row = hdr_row + 1 + i
        ws.cell(row=row, column=1, value=i + 1).alignment = ctr
        ws.cell(row=row, column=1).border = brd
        ws.cell(row=row, column=1).font = Font(name="Arial", size=10)

        nm_cell = ws.cell(row=row, column=2, value=name)
        nm_cell.alignment = rgt
        nm_cell.border = brd
        nm_cell.font = Font(name="Arial", size=10)

        row_vals = matrix[i] if i < len(matrix) else []
        for j in range(len(dates)):
            col = 3 + j
            tok = row_vals[j] if j < len(row_vals) else "—"
            c = ws.cell(row=row, column=col, value=tok)
            c.alignment = ctr
            c.border = brd
            c.font = Font(name="Arial", size=9, bold=True)
            c.fill = _cell_fill_for_day_token(tok, pfill, afill, lfill, blank_fill)

        p, a, l_, pct = summaries.get(name, (0, 0, 0, 0.0))
        for k, val in enumerate([p, a, l_, pct]):
            c = ws.cell(row=row, column=base_col + k, value=val)
            c.alignment = ctr
            c.border = brd
            c.font = Font(name="Arial", size=10)

        ws.row_dimensions[row].height = 20

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 36
    for j in range(len(dates)):
        col_letter = get_column_letter(3 + j)
        ws.column_dimensions[col_letter].width = 5.5
    for k in range(4):
        col_letter = get_column_letter(base_col + k)
        ws.column_dimensions[col_letter].width = 10

    legend_row = hdr_row + len(student_names) + 2
    ws.cell(row=legend_row, column=1, value="مفتاح: ح = حاضر | غ = غائب | م = متأخر | — = لا سجل")
    ws.merge_cells(
        start_row=legend_row,
        start_column=1,
        end_row=legend_row,
        end_column=min(6, 2 + len(dates)),
    )
    ws.cell(row=legend_row, column=1).font = Font(name="Arial", italic=True, size=10, color="555555")
    ws.cell(row=legend_row, column=1).alignment = rgt

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def daterange_inclusive(start: date, end: date) -> List[date]:
    if end < start:
        start, end = end, start
    out = []
    cur = start
    while cur <= end:
        out.append(cur)
        cur += timedelta(days=1)
    return out


def month_dates(year: int, month: int) -> List[date]:
    last = monthrange(year, month)[1]
    return [date(year, month, d) for d in range(1, last + 1)]
