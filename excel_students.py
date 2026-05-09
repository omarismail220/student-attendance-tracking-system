"""
Excel helpers for RTL student roster template (openpyxl) and safe import parsing.

Template layout (contracts with ``parse_student_workbook``):

* Row 1: merged title cells with group label and time (informational — not imported).
* Row 2: column header; cell A2 must equal ``STUDENTS_COL_HEADER``.
* Row 3 onward: student names in column A until the sheet ends; blank cells skipped.
"""

from __future__ import annotations

import io
import re
from typing import Iterable, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# Fixed header literal — import rejects files whose A2 does not normalize to this text.
STUDENTS_COL_HEADER = "أسماء الطلاب"

_WS_RE = re.compile(r"\s+")


def _norm_header(s: str) -> str:
    return _WS_RE.sub(" ", (s or "").strip())


def normalize_student_name(cell_value) -> str | None:
    """Turn a worksheet cell value into a trimmed Arabic-friendly string or None."""
    if cell_value is None:
        return None
    if isinstance(cell_value, float) and cell_value.is_integer():
        cell_value = str(int(cell_value))
    text = _norm_header(str(cell_value))
    return text if text else None


def dedupe_preserve_order(names: Iterable[str]) -> List[str]:
    """Drop duplicates while preserving first-seen order (Arabic-sensitive, exact match)."""
    seen = set()
    out: List[str] = []
    for n in names:
        if n not in seen:
            seen.add(n)
            out.append(n)
    return out


def build_student_template_roster_workbook(
    group_id: str, label: str, time_: str, students: List[str]
) -> Tuple[bytes, str]:
    """
    Build workbook bytes for GET /api/template/<group_id>.

    Returns ``(raw_xlsx_bytes, suggested_filename)``.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "طلاب"

    ws.sheet_view.rightToLeft = True
    merged_last_row = 1
    ws.merge_cells(start_row=1, start_column=1, end_row=merged_last_row, end_column=2)
    ws["A1"] = f"{label} — {group_id} | الموعد: {time_}"
    title_font = Font(name="Arial", bold=True, size=13, color="1a3c5e")
    title_fill = PatternFill("solid", start_color="e8f0fe")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="cccccc")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["A1"].font = title_font
    ws["A1"].alignment = ctr
    ws["A1"].fill = title_fill
    ws["A1"].border = brd
    ws.row_dimensions[1].height = 32

    hdr_row = 2
    hf = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    hfill = PatternFill("solid", start_color="1a3a5c")
    hdr_align = Alignment(horizontal="center", vertical="center")
    hc = ws.cell(row=hdr_row, column=1, value=STUDENTS_COL_HEADER)
    hc.font = hf
    hc.fill = hfill
    hc.alignment = hdr_align
    hc.border = brd
    ws.row_dimensions[hdr_row].height = 26

    body_font = Font(name="Arial", size=11)
    name_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    row0 = hdr_row + 1
    for offset, name in enumerate(students):
        r = row0 + offset
        c = ws.cell(row=r, column=1, value=name)
        c.font = body_font
        c.alignment = name_align
        c.border = brd

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 8

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"students_template_{group_id}.xlsx"
    return buf.getvalue(), fname


def parse_student_workbook(
    file_bytes: bytes,
    *,
    max_bytes: int,
    max_student_rows: int,
) -> Tuple[List[str], str | None]:
    """
    Validate and parse uploaded roster workbook.

    :param file_bytes: Raw .xlsx content.
    :param max_bytes: Hard cap on uploaded size (checked by caller too).
    :param max_student_rows: Safety ceiling for rows scanned under A column.
    :returns: (names, error_ar_message or None).

    Raises nothing — corruption returns ``([], "…")``.
    """
    if len(file_bytes) > max_bytes:
        return [], "حجم الملف أكبر من المسموح"

    try:
        bio = io.BytesIO(file_bytes)
        wb = load_workbook(bio, read_only=True, data_only=True)
    except Exception:
        return [], "ملف Excel تالف أو غير صالح"

    try:
        ws = wb.active
        if ws is None:
            return [], "الملف لا يحتوي ورقة عمل"

        a2 = ws["A2"].value
        if _norm_header(str(a2 or "")) != STUDENTS_COL_HEADER:
            return [], (
                "تنسيق الملف غير صحيح: يجب أن يحتوي الصف الثاني على عنوان العمود "
                f"«{STUDENTS_COL_HEADER}» في الخلية A2 دون تغيير"
            )

        names: List[str] = []
        start_row = 3
        end_row = min(start_row + max_student_rows, ws.max_row + 1)
        for row in range(start_row, end_row):
            val = normalize_student_name(ws.cell(row=row, column=1).value)
            if val:
                names.append(val)

        if not names:
            return [], "لم يتم العثور على أي أسماء طلاب تحت عنوان العمود"

        return dedupe_preserve_order(names), None
    finally:
        try:
            wb.close()
        except Exception:
            pass
