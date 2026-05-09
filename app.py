from __future__ import annotations

import io
import os
import re
import threading
from datetime import datetime
from pathlib import Path

from flask import Flask, jsonify, request, send_file, send_from_directory
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from werkzeug.exceptions import RequestEntityTooLarge
from werkzeug.utils import secure_filename

from excel_students import build_student_template_roster_workbook, parse_student_workbook
from storage import (
    load_attendance_disk,
    load_groups_disk,
    prune_attendance_for_group,
    save_attendance,
    save_groups,
)

# ── Flask app ─────────────────────────────────────────────────────────────────
_app_root = Path(__file__).resolve().parent
app = Flask(__name__)

# Reject oversized uploads early (multipart body). Tune for PythonAnywhere if needed.
MAX_UPLOAD_BYTES = 2 * 1024 * 1024  # 2 MiB — plenty for roster lists
MAX_STUDENT_IMPORT_ROWS = 2000

app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_BYTES

# In-memory stores synced with JSON on disk — load once at startup, persist after writes.
_groups_lock = threading.Lock()
_groups: dict = {}
attendance_store: dict = {}

# ── CORS ───────────────────────────────────────────────────────────────────────
CORS(
    app,
    origins=[
        "https://student-attendance-tracking-system-3un4pgeat.vercel.app",
        "https://omarismail220.pythonanywhere.com",
        "http://127.0.0.1:5050",
        "http://localhost:5050",
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "null",
        re.compile(r"^https://[a-zA-Z0-9.-]+\.vercel\.app$"),
    ],
)


def bootstrap_data_stores() -> None:
    """
    Load persisted groups + attendance into process memory.

    Groups file is created on first run under ``data/groups.json`` from defaults in
    ``storage.DEFAULT_GROUPS``. Attendance uses ``data/attendance.json`` or starts empty.
    """
    global _groups, attendance_store
    _groups = load_groups_disk(_app_root)
    attendance_store = load_attendance_disk(_app_root)
    for gid, meta in _groups.items():
        if gid not in attendance_store or not isinstance(attendance_store.get(gid), dict):
            attendance_store[gid] = {}
        prune_attendance_for_group(
            gid,
            set(meta.get("students", [])),
            attendance_store,
        )
    save_attendance(attendance_store, _app_root)


bootstrap_data_stores()

PA_SAMPLE_XLSX_DEFAULT = "/home/omarismail220/sample_attendance_A1.xlsx"


def resolve_sample_xlsx_path():
    """Resolve optional sample attendance file path (deployment-specific)."""
    env = os.environ.get("ATTENDANCE_SAMPLE_XLSX", "").strip()
    if env and os.path.isfile(env):
        return env
    local = os.path.join(_app_root, "sample_attendance_A1.xlsx")
    if os.path.isfile(local):
        return local
    if os.path.isfile(PA_SAMPLE_XLSX_DEFAULT):
        return PA_SAMPLE_XLSX_DEFAULT
    return None


@app.route("/")
def index():
    """Serve static SPA HTML from project root."""
    return send_from_directory(str(_app_root), "index.html")


@app.route("/api/info", methods=["GET"])
def api_info():
    """
    Lightweight health + metadata endpoint for dashboards and frontend status badge.

    Returns group count from persisted store so it reflects roster JSON, not literals.
    """
    sample = resolve_sample_xlsx_path()
    return jsonify(
        {
            "status": "ok",
            "service": "student-attendance-api",
            "version": "2.0.0",
            "groups_count": len(_groups),
            "server_time": datetime.now().isoformat(timespec="seconds"),
            "sample_xlsx_available": bool(sample),
        }
    )


@app.route("/api/sample-attendance-xlsx", methods=["GET"])
def download_sample_attendance_xlsx():
    """Download prebuilt sample attendance Excel (existing feature)."""
    path = resolve_sample_xlsx_path()
    if not path:
        return (
            jsonify(
                {"error": "not_found", "message": "ملف النموذج غير موجود على الخادم"}
            ),
            404,
        )
    return send_file(
        path,
        as_attachment=True,
        download_name="sample_attendance_A1.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/groups", methods=["GET"])
def get_groups():
    """
    List teaching groups — each entry exposes id/label/time and live student counts.

    Data source: persisted ``groups.json`` loaded at startup and updated by Excel import.
    """
    list_res = []
    with _groups_lock:
        for k, v in _groups.items():
            list_res.append(
                {
                    "id": k,
                    "label": v["label"],
                    "time": v["time"],
                    "count": len(v["students"]),
                }
            )
    return jsonify(list_res)


@app.route("/api/students/<group_id>", methods=["GET"])
def get_students(group_id):
    """Return roster + metadata + current in-memory attendance map for UI."""
    with _groups_lock:
        if group_id not in _groups:
            return jsonify({"error": "not_found", "message": "المجموعة غير موجودة"}), 404
        g = _groups[group_id]
        snap = {"label": g["label"], "time": g["time"], "students": list(g["students"])}
        att_block = attendance_store.get(group_id, {})
    snap["attendance"] = att_block if isinstance(att_block, dict) else {}
    return jsonify(snap)


@app.route("/api/template/<group_id>", methods=["GET"])
def download_student_template(group_id):
    """
    Download an RTL Arabic .xlsx template for editing student names offline.

    Layout: row 1 merged title with group label + slot time; row 2 header styled;
    rows 3+ prefilled student names in column A. Import endpoint expects unchanged A2 header.
    """
    with _groups_lock:
        if group_id not in _groups:
            return jsonify({"error": "not_found", "message": "المجموعة غير موجودة"}), 404
        g = _groups[group_id]
        label = g["label"]
        tm = g["time"]
        students = list(g["students"])
    buf, fname = build_student_template_roster_workbook(group_id, label, tm, students)
    return send_file(
        io.BytesIO(buf),
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/import-students/<group_id>", methods=["POST"])
def import_students(group_id):
    """
    Replace a group's roster from an uploaded .xlsx (multipart field ``file``).

    Validates extension, UTF-8–safe Arabic text, header cell A2 contract, parses column A from
    row 3 onward, trims blanks, strips duplicates deterministically (first occurrence wins).

    Persist: writes atomically to ``groups.json``, prunes unknown student keys across all
    saved attendance dates, then persists ``attendance.json``.
    """
    with _groups_lock:
        if group_id not in _groups:
            return jsonify({"error": "not_found", "message": "المجموعة غير موجودة"}), 404

    uf = request.files.get("file")
    if uf is None or uf.filename == "":
        return (
            jsonify({"success": False, "error": "missing_file", "message": "لم يُرفع أي ملف"}),
            400,
        )

    safe = secure_filename(uf.filename)
    if not safe.lower().endswith(".xlsx"):
        return (
            jsonify(
                {
                    "success": False,
                    "error": "bad_extension",
                    "message": "يُسمح فقط بملفات امتداد .xlsx",
                }
            ),
            415,
        )

    raw = uf.read()
    if len(raw) == 0:
        return (
            jsonify({"success": False, "error": "empty_file", "message": "الملف فارغ"}),
            400,
        )
    if len(raw) > MAX_UPLOAD_BYTES:
        return (
            jsonify({"success": False, "error": "file_too_large", "message": "حجم الملف كبير جداً"}),
            413,
        )

    names, parse_err = parse_student_workbook(
        raw, max_bytes=MAX_UPLOAD_BYTES, max_student_rows=MAX_STUDENT_IMPORT_ROWS
    )
    if parse_err:
        return (
            jsonify({"success": False, "error": "invalid_template", "message": parse_err}),
            422,
        )

    with _groups_lock:
        gp = dict(_groups[group_id])
        gp["students"] = names
        _groups[group_id] = gp
        save_groups(_groups, _app_root)

        prune_attendance_for_group(group_id, set(names), attendance_store)
        save_attendance(attendance_store, _app_root)

    return jsonify(
        {"success": True, "group_id": group_id, "imported_count": len(names)}
    )


@app.route("/api/attendance/<group_id>", methods=["GET"])
def get_attendance_history(group_id):
    """Return persisted day→records map for a group."""
    with _groups_lock:
        if group_id not in _groups:
            return jsonify({"error": "not_found"}), 404
    raw = attendance_store.get(group_id, {})
    return jsonify(raw)


@app.route("/api/attendance", methods=["POST"])
def save_attendance_route():
    """
    Persist UI attendance grid for ``group_id``, ``date``, and ``records``.

    Mirrors prior JSON contract — now also appends durability by writing attendance.json.
    """
    data = request.json or {}
    gid = data.get("group_id")
    records = data.get("records", {})
    date = data.get("date", datetime.now().strftime("%Y-%m-%d"))
    if not isinstance(records, dict):
        records = {}

    with _groups_lock:
        if gid not in _groups:
            return jsonify({"error": "not_found"}), 404

    attendance_store.setdefault(gid, {})
    attendance_store[gid][date] = records
    save_attendance(attendance_store, _app_root)
    return jsonify({"status": "saved"})


@app.route("/api/export/<group_id>", methods=["GET"])
def export_attendance(group_id):
    """Generate RTL attendance workbook for Excel download — roster from persisted JSON."""
    date = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))

    with _groups_lock:
        if group_id not in _groups:
            return jsonify({"error": "not found"}), 404
        group = dict(_groups[group_id])

    records = attendance_store.get(group_id, {}).get(date, {})

    wb = Workbook()
    ws = wb.active
    ws.title = "كشف الحضور"
    ws.sheet_view.rightToLeft = True

    hf = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    hfill = PatternFill("solid", start_color="1a3a5c")
    pfill = PatternFill("solid", start_color="d4edda")
    afill = PatternFill("solid", start_color="fee2e2")
    ctr = Alignment(horizontal="center", vertical="center")
    rgt = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="cccccc")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:D1")
    ws["A1"] = f"كشف حضور - {group['label']} | {group['time']} | تاريخ: {date}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1a3c5e")
    ws["A1"].alignment = ctr
    ws["A1"].fill = PatternFill("solid", start_color="e8f0fe")
    ws.row_dimensions[1].height = 30

    for col, h in enumerate(["م", "اسم الطالب", "الحضور", "ملاحظات"], 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hf
        c.fill = hfill
        c.alignment = ctr
        c.border = brd

    ws.row_dimensions[2].height = 25

    def is_present(raw):
        return raw in ("present", "حاضر", "متاخر")

    present_count = 0

    student_list = list(group["students"])
    for i, name in enumerate(student_list, 1):
        row = i + 2
        raw_s = records.get(name, "absent")
        status = "present" if is_present(raw_s) else "absent"
        if status == "present":
            present_count += 1

        fill = pfill if status == "present" else afill
        label = "متأخر ⏰" if raw_s == "متاخر" else ("حاضر ✔" if status == "present" else "غائب ✘")

        for col, (val, aln) in enumerate(zip([i, name, label, ""], [ctr, rgt, ctr, ctr]), 1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = aln
            c.border = brd
            c.font = Font(name="Arial", size=11)
            if col in (1, 3):
                c.fill = fill
        ws.row_dimensions[row].height = 22

    total = len(student_list)
    sr = total + 3
    ws.merge_cells(f"A{sr}:B{sr}")
    ws[f"A{sr}"] = f"إجمالي الحاضرين: {present_count} / {total}"
    ws[f"A{sr}"].font = Font(name="Arial", bold=True, size=11, color="155724")
    ws[f"A{sr}"].fill = pfill
    ws[f"A{sr}"].alignment = ctr
    ws[f"A{sr}"].border = brd
    ws.merge_cells(f"C{sr}:D{sr}")
    ws[f"C{sr}"] = f"إجمالي الغائبين: {total - present_count}"
    ws[f"C{sr}"].font = Font(name="Arial", bold=True, size=11, color="721c24")
    ws[f"C{sr}"].fill = afill
    ws[f"C{sr}"].alignment = ctr
    ws[f"C{sr}"].border = brd

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"attendance_{group_id}_{date}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


API_ENDPOINTS = [
    "GET /api",
    "GET /api/info",
    "GET /api/groups",
    "GET /api/students/<group_id>",
    "POST /api/attendance",
    "GET /api/attendance/<group_id>",
    "GET /api/export/<group_id>",
    "GET /api/sample-attendance-xlsx",
    "GET /api/template/<group_id>",
    "POST /api/import-students/<group_id>",
    "GET /",
]


@app.route("/api", methods=["GET"])
def api_index():
    return jsonify({"service": "student-attendance-api", "endpoints": API_ENDPOINTS}), 200


@app.errorhandler(RequestEntityTooLarge)
def handle_too_large(_):
    """Uniform JSON for bodies exceeding MAX_CONTENT_LENGTH."""
    return jsonify({"success": False, "error": "payload_too_large", "message": "طلب أكبر من الحد الأقصى"}), 413


@app.errorhandler(404)
def handle_404(exc):
    if request.path.startswith("/api"):
        return (
            jsonify(
                {"error": "not_found", "path": request.path, "endpoints": API_ENDPOINTS},
            ),
            404,
        )
    body = (
        "<!DOCTYPE html><html><body dir=\"rtl\" style=\"font-family:sans-serif;padding:2rem;\">"
        f"<h1>404</h1><p>المسار غير موجود: <code>{request.path}</code></p>"
        '<p><a href="/">الصفحة الرئيسية</a> — <a href="/api/info">/api/info</a></p></body></html>'
    )
    return body, 404


if __name__ == "__main__":
    app.run(debug=False, port=5050)
